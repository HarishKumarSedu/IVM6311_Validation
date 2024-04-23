import openpyxl
from pathlib import Path
import json
import sys
from time import sleep


# xlsx_file = Path(r"C:\Users\galfieri\Documents\projects\Python\Test\Excel to json", "IVM6303_rev3_regmap.xlsx")
# xlsx_file = Path(r"C:\Users\galfieri\Documents\projects\Python\GUI from excel", "Tabit_rev1_regmap.xlsx")
# xlsx_file = Path("./", "Tabit_rev1_regmap.xlsx")
xlsx_file = Path("./", "raw_data/IVM6311_rev0_regmap.xlsx")
regmap_list = []


wb_obj = openpyxl.load_workbook(xlsx_file, data_only=True)
wb_obj.active = 0
ws = wb_obj.active
print(ws)

adr_col = 1
msb_col = 6
attr_col = 15
page = 0
sections = []
sec_dict = dict()
sec_dict["List_Tab"] = []

for ind_sheet in range(len(wb_obj.get_sheet_names())):
    wb_obj.active = ind_sheet
    ws = wb_obj.active
    page_title = ws.title[3:4]
    if page_title == str(page):
        addr = 0
        addr_hex = '{0:0{1}X}'.format(addr,2)
        for i in range(1, ws.max_row):
            cell_contents = ws.cell(i, adr_col).value
            if cell_contents == addr_hex:
                bit = 7
                attribute = str(ws.cell(i, attr_col).value)
                index = 0
                for j in range(msb_col, msb_col + 8):
                    field_name = str(ws.cell(i, j).value)
                    if addr_hex == 'EF':
                        stop = True
                    if (field_name != "(reserved)") and (field_name != "None") and (field_name != "(spare)"):
                        # print('********',ws.cell(i,14).value)
                        
                        register_name = ws.cell(i,14).value
                        
                        len_info = field_name.find("[")
                        len_info2 = field_name.find("]")
                        if (len_info == -1) or ((len_info2 - len_info) <= 2):
                            len= 1
                        else:
                            check = []
                            for c in range(len_info+1, len_info+3):
                                check.append(field_name[c].isnumeric())
                            msb_index1 = len_info+1
                            msb_index2 = msb_index1 + 1
                            lsb_index1 = len_info+3
                            lsb_index2 = lsb_index1 + 1
                            if check[1] == True:
                                msb_index2 = msb_index2 + 1
                                lsb_index1 = lsb_index1 + 1
                                lsb_index2 = lsb_index2 + 1
                            check = []
                            for c in range(lsb_index1, lsb_index1 + 2 ):
                                check.append(field_name[c].isnumeric())
                            if check[1] == True:
                                lsb_index2 = lsb_index2 + 1

                            # len = int(field_name[len_info+1]) - int(field_name[len_info + 3]) + 1
                            len = int(field_name[msb_index1:msb_index2]) - int(field_name[lsb_index1:lsb_index2]) + 1
                        option = (str(ws.cell(i+1, j).value)).replace("\n", "\t")
                        attr_field = attribute[index]
                        lsb = bit - len + 1
                        match = ""
                        
                        for regfield in regmap_list:
                            name = regfield["Name"][:regfield["Name"].index("[")] if "[" in regfield["Name"] else "-1"
                            field_only = field_name[:field_name.index("[")] if "[" in field_name else "-2"
                            # print(name, field_only)
                            if name == field_only:
                                if regfield["Match"] != "":
                                    match += (" " + regfield["REG"])
                                    regfield["Match"] += (" " + str(addr_hex))
                                else:
                                    match = regfield["REG"]
                                    regfield["Match"] = str(addr_hex)




                        # print("PAGE " + str(page) +  " REG " + addr_hex  + " Bit " + str(bit) + " LSB " + str(lsb) + " " + field_name + " Length " + str(len) +
                            #   " Option: " + option + " Attribute: " + attr_field + " Match: " + match)

                        regmap = {
                            "PAGE" : page,
                            "REG" : addr_hex,
                            "REGNAME":register_name,
                            "POS" : lsb,
                            "Name": field_name,
                            "Length": len,
                            "Option": option,
                            "Attribute" : attr_field,
                            "Match" : match
                        }
                        # regmap_dict['']=regmap
                        regmap_list.append(regmap)

                    index += 1
                    bit -=1
                addr +=1
                addr_hex = '{0:0{1}X}'.format(addr,2)
            else:
                section_name = str(cell_contents)
                if (section_name not in "None") and (section_name not in "Hex"):
                    if addr_hex != "00":
                        sections[-1].append("{:02X}".format(addr - 1))
                    sections.append([section_name, page, addr_hex])
        page += 1

print(sections)
for sec in sections:
    field = dict()
    field["Name"] = sec[0]
    p = sec[1]
    addr = sec[2]
    try:
        if sec[3] in sec:
            addr += (" - " + sec[3])
    except:
        pass
    field["Regs"] = [[p, addr]]
    field["Fields"] = []
    sec_dict["List_Tab"].append(field)

regmap_write = {"RegMap" : regmap_list}
with open("regmap_IVM6311.json", "w") as outfile:
    json.dump(regmap_write, outfile, indent = 4)

with open("sections_IVM6311.json", "w") as outfile:
    json.dump(sec_dict, outfile, indent = 4)